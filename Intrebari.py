import tkinter as tk

import openpyxl



def get_answer(question, file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] == question:
            return row[1]
    return "Nu am găsit un răspuns pentru această întrebare."

def on_submit():
    question = entry.get()
    if question:
        if question.lower() == 'stop':
            root.quit()
        if question in questions:
            answer = get_answer(question, excel_file_path)
            result_label.config(text="Răspuns: " '\n' + answer)
        else:
            result_label.config(text="Întrebare invalidă.")

root = tk.Tk()
root.title("Pufulete")

excel_file_path = 'C:\\Users\\rares\\Desktop\\Hackathon text\\rasp.xlsx'

wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active
questions = [row[0] for row in sheet.iter_rows(values_only=True)]

question_label = tk.Label(root, text="Intrebare:")
question_label.pack()

entry = tk.Entry(root)
entry.pack()

submit_button = tk.Button(root, text="Trimite", command=on_submit)
submit_button.pack()

result_label = tk.Label(root, text="")
result_label.pack()

root.mainloop()
